import matplotlib.pyplot as plt
import matplotlib.image as mpimg
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook


def extract_node_coordinates(image_path, excel_file_path, num_nodes):
    """
    Interactive tool to extract node coordinates by clicking on an image.
    
    Parameters
    ----------
    image_path : str
        Path to the reference image
    excel_file_path : str
        Path to the Excel file where coordinates will be added as a new sheet
    num_nodes : int
        Expected number of nodes to click
    """
    
    # Load the image
    img = mpimg.imread(image_path)
    
    # Create figure and display image
    fig, ax = plt.subplots(figsize=(16, 12))
    ax.imshow(img)
    ax.set_title(f'Click on each intersection in order (1 to {num_nodes})\nPress any key when done', 
                 fontsize=14, fontweight='bold')
    ax.axis('on')  # Keep axes to see coordinates
    
    # Storage for coordinates
    coordinates = {}
    current_node = 1
    
    def onclick(event):
        nonlocal current_node
        
        if event.inaxes != ax:
            return
            
        if current_node <= num_nodes:
            # Store coordinates (flip y-axis since image coordinates are top-down)
            x, y = event.xdata, event.ydata
            coordinates[current_node] = (x, img.shape[0] - y)  # Flip y for standard coordinate system
            
            # Add a marker and label
            ax.plot(x, y, 'ro', markersize=8)
            ax.annotate(str(current_node), (x, y), xytext=(5, 5), 
                       textcoords='offset points', fontsize=10, fontweight='bold',
                       bbox=dict(boxstyle='round,pad=0.3', facecolor='yellow', alpha=0.8))
            
            print(f"Node {current_node}: ({x:.1f}, {img.shape[0] - y:.1f})")
            current_node += 1
            
            if current_node > num_nodes:
                ax.set_title(f'All {num_nodes} nodes captured! Press any key to save and exit', 
                           fontsize=14, fontweight='bold', color='green')
            else:
                ax.set_title(f'Click on intersection {current_node} (of {num_nodes})', 
                           fontsize=14, fontweight='bold')
            
            plt.draw()
    
    def onkey(event):
        if coordinates:
            # Save coordinates to Excel file as a new sheet
            save_coordinates_to_excel(coordinates, excel_file_path)
            plt.close()
    
    # Connect event handlers
    fig.canvas.mpl_connect('button_press_event', onclick)
    fig.canvas.mpl_connect('key_press_event', onkey)
    
    # Instructions
    print(f"Instructions:")
    print(f"1. Click on each intersection in numerical order (1 to {num_nodes})")
    print(f"2. The script will mark each clicked point with a red dot and number")
    print(f"3. Press any key when you're done to save coordinates")
    print(f"4. Coordinates will be added to: {excel_file_path}")
    print(f"\nStarting with node 1...")
    
    plt.show()
    
    return coordinates


def save_coordinates_to_excel(coordinates, excel_file_path):
    """
    Save coordinates to an Excel file as a new "Coordinates" sheet.
    Preserves all existing sheets with their formatting, formulas, and colors.
    
    Parameters
    ----------
    coordinates : dict
        Dictionary with node IDs as keys and (x, y) tuples as values
    excel_file_path : str
        Path to the Excel file
    """
    
    # Create a DataFrame from coordinates
    coord_data = []
    for node_id in sorted(coordinates.keys()):
        x, y = coordinates[node_id]
        coord_data.append({
            'Node': node_id,
            'X': x,
            'Y': y
        })
    
    coords_df = pd.DataFrame(coord_data)
    
    excel_file = Path(excel_file_path)
    
    # Load existing workbook or create new one
    if excel_file.exists():
        try:
            workbook = load_workbook(excel_file_path)
        except Exception as e:
            print(f"Warning: Could not load existing workbook ({e}). Creating new one.")
            workbook = Workbook()
    else:
        workbook = Workbook()
    
    # Remove the "Coordinates" sheet if it already exists
    if 'Coordinates' in workbook.sheetnames:
        workbook.remove(workbook['Coordinates'])
    
    # Create a new "Coordinates" sheet
    coords_sheet = workbook.create_sheet('Coordinates')
    
    # Add headers
    coords_sheet['A1'] = 'Node'
    coords_sheet['B1'] = 'X'
    coords_sheet['C1'] = 'Y'
    
    # Add data
    for i, (node_id, (x, y)) in enumerate(sorted(coordinates.items()), start=2):
        coords_sheet[f'A{i}'] = node_id
        coords_sheet[f'B{i}'] = x
        coords_sheet[f'C{i}'] = y
    
    # Save the workbook (preserves all existing sheets with their formatting)
    workbook.save(excel_file_path)
    
    print(f"\nCoordinates saved to sheet 'Coordinates' in: {excel_file}")
    print(f"Captured {len(coordinates)} nodes")
    print("All existing sheets and their formatting have been preserved.")


if __name__ == "__main__":

    # Paths
    image_path = "../data/37-intersection map.png"
    excel_file_path = "../data/37-intersection map.xlsx"
    num_nodes = 38  # Including ending point (duplicate of starting point)
    
    # Extract coordinates
    coords = extract_node_coordinates(image_path, excel_file_path, num_nodes=num_nodes)
    
    if coords:
        print(f"\nExtracted coordinates for {len(coords)} nodes:")
        for node_id in sorted(coords.keys()):
            x, y = coords[node_id]
            print(f"  {node_id}: ({x:.1f}, {y:.1f})")
